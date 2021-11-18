import os
import argparse
import numpy as np
from pandas import Series
from itertools import chain
import openpyxl as xl
from datetime import datetime

from mikeio import Dfs0, read
from mikeio.eum import ItemInfo, EUMType, EUMUnit

parser = argparse.ArgumentParser()
# Switch
parser.add_argument("-m","--mode",
                    help="Run in read (r) or write (w) mode",
                    choices = ['r','w'],
                    type=str)
# Universal
parser.add_argument("-wb","--workbook",
                    help="MikeBasin Excel workbook with input data",
                    type=str)

parser.add_argument("-ws","--datasheet",
                    help="Data sheet name",
                    type=str)
# Write dfs0 args
parser.add_argument("-r","--data_range",
                    default = "B137:EX2788",
                    help="Range of data including headers",
                    type=str)

parser.add_argument("-rf","--refsheet",
                    default = "Reference",
                     help="Node name (col 1) and DHI ID (col 2) table",
                     type=str)

parser.add_argument("-rr","--refrow",
                    default = 10,
                    help="First row of node and ID data. Excludes header",
                    type=int)
parser.add_argument("-od","--outdir",
                     default = "data",
                     help = "Directory to place exported dfs0 files."+\
                         "Defaults to ./data and creates directory if it does not exist",
                     type=str)
# Read dfs0 args
parser.add_argument("-hr","--header_range",
                    help="Range of data header",
                    type=str)

parser.add_argument("-ss","--scenario_sheet",
                    help = "Table listing scenarios to be loaded from results",
                    default = "Master",
                    type = str)
parser.add_argument("-sr","--scenario_range",
                    help = "Cell range describing where scenario instructions are",
                    type = str)
parser.add_argument("-t","--temp_workbook",
                    default = f"dfs0_{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx",
                    help = "Full path to temporary .xlsx file holding model results.",
                    type = str)

args = parser.parse_args()

#Universal
mode = args.mode
workbook = args.workbook
datasheet = args.datasheet

# Write dfs0
cell_range_data = args.data_range
refsheet = args.refsheet
refrow = args.refrow
outdir = args.outdir

# Read dfs0
hdr_range = args.header_range
scenario_sheet = args.scenario_sheet
scenario_range = args.scenario_range
ftemppath = args.temp_workbook

#--------------------------------
# Write dfs0 files for Mike Basin
#--------------------------------
def empty():
    yield from ()

def excel2dfs0():
    # Create output directory if it does not exist
    if not os.path.exists(outdir):
        print('\t',f'Creating output dfs0 folder {os.path.join(os.getcwd(),outdir)}')
        os.mkdir(outdir)

    wb = xl.load_workbook(workbook,read_only=True,data_only=True, keep_vba= False)
    # Map node name to DHI ID
    ref_sheet = wb[refsheet]
    node_dict={}
    for r in ref_sheet.iter_rows(min_col=1,max_col=2,min_row=refrow):
        if r[0].value:
            node_dict.update({r[1].value:r[0].value})
        else:
            break

    # Load sheet and determine table indices
    sheet = wb[datasheet]
    col_min_data,row_min_data,col_max_data,row_max_data=\
    xl.utils.cell.range_boundaries(cell_range_data)

    #-----------------------
    #   Gather data for export
    #----------------------

    ## Read Only Method
    datad={}
    for i,r in enumerate(sheet.iter_rows(min_col=col_min_data,
                                        max_col=col_max_data,
                                        min_row=row_min_data,
                                        max_row=row_max_data)):
        if i == 0:
            units = [c.value.strip('[]').strip() for c in r][1:]
        elif i == 1:
            items = [c.value for c in r][1:]
        elif i == 2:
            LoadData = [c.value for c in r][1:]
        else:
            record = [c.value for c in r ]
            # datad.update({record[0]:  record[1:]  } )
            datad.update({record[0]: (n for n in record[1:] ) } )


    #------------------------------------
    # Write dfs0 to folder. No pandas use
    # -------------------------------------
    # Transpose
    d = {}
    for ts_idx,row in enumerate(datad.values()):
        for item_idx,cell in enumerate(row):
            val_list = d.get(items[item_idx])
            if not val_list:
                val_list =  empty()
            val_list = chain(val_list,[cell])
            d.update({items[item_idx]: val_list})

    # Keys: Item Name, values: tuple (load data, units)
    item_dict = dict(zip(items,zip(LoadData,units)))
    # Timestamps. Just need to do this once
    ts = list(datad.keys())

    for item,item_meta in item_dict.items():
        if item_meta[0]:
            unit = item_meta[1]
            # Get DHI ID and map to name
            dhi_id = item.split("|")[0]
            fname = f"{node_dict[dhi_id]}.dfs0"

            # Rename output if item is a return flow
            if "return" in item.lower():
                fname = fname.replace(".dfs0","_rf.dfs0")

            # Initialize dfs0 object
            dfs = Dfs0()
            if unit == 'cfs':
                mike_type = EUMType.Water_Flow
                mike_unit = EUMUnit.feet_pow_3_per_sec
            elif (not unit) & ('Return' in item):
                mike_type = EUMType.Return_Flow_Fraction
                mike_unit = EUMType.Return_Flow_Fraction.units[0]

            # Create Mike Item object
            mike_item = ItemInfo(item,mike_type,unit=mike_unit)

            # Create dfs0
            print(f"Writing {item}")
            data = np.array(list(d[item]))
            data = data.astype(float)

            dfs.write(filename=os.path.join(outdir,fname),
                       items = [mike_item],
                       datetimes= ts,
                       data= [data])


    # Close up workbook
    wb.close()

#--------------------------------
# Read Mike Basin results into excel file
#--------------------------------
def dfs02excel():
    """
    workbook = r"CHCBM-FedDiv_All-32BIT-v03_test.xlsm"  #Name of Workbook
    datasheet = "BC1 Data"
    hdr_range = "A47:BX52"

    scenario_sheet = "Master"
    scenario_range = "A3:H8"

    """
    # Load workbook
    wb = xl.load_workbook(workbook,read_only=True,data_only=True, keep_vba=False)

    # Lookup scenarios to load
    ws_scenario = wb[scenario_sheet]

    col_min_data,row_min_data,col_max_data,row_max_data=\
    xl.utils.cell.range_boundaries(scenario_range)

    iterator = ws_scenario.iter_rows(min_col=col_min_data,max_col=col_max_data,
                                     min_row=row_min_data,max_row=row_max_data)

    # Manage Scenarios
    l=[]

    for i,r in enumerate(iterator):
        arr=[c.value for c in r]
        l.append(arr)

    # Create dictionary for requested scenarios.
    # This is heavily dependent on the table structure and could easily break
    scenario_dict = {item[0]:{'Filename':item[1],'Load':item[-1]} for item in l if (item[0]!='Scenario') and (item[1])}

    # Read data entry sheet
    ws = wb[datasheet]

    # Header Dictionary for excel sheet indices.
    col_min_data,row_min_data,col_max_data,row_max_data=\
        xl.utils.cell.range_boundaries(hdr_range)
    iterator = ws.iter_rows(min_col=col_min_data,max_col=col_min_data,
                                      min_row=row_min_data,max_row=row_max_data)
    hdr_dict = {c.value:c.row for r in iterator for c in r}

    # Read data header into dataframe. May not need this??
    # col_min_data,row_min_data,col_max_data,row_max_data=\
    # xl.utils.cell.range_boundaries(hdr_range)
    # iterator = ws.iter_rows(min_col=col_min_data,max_col=col_max_data,
    #                                  min_row=row_min_data,max_row=row_max_data)
    # data = {i: {j: c.value for j,c in enumerate(r)} for i,r in enumerate(iterator)}
    # data = DataFrame(data)
    # df2=df.T.set_index(0).T.pad().set_index('Scenario')


    # Determine appropriate data range for the scenario
    # Pull expected scenario header row
    scenario_row = hdr_dict['Scenario']
    iterator = ws.iter_rows(min_col=1, max_col=ws.max_column,
                            min_row=scenario_row, max_row=scenario_row)
    r = next(iterator)

    # Helper function
    def get_to_column(s):
        '''Takes a series of excel columns, converts to index, shifts the series,
        subtracts one to get the last column of the previous cell with a string,
        converts back to a column letter'''
        return s.apply(xl.utils.cell.column_index_from_string)\
                .shift(-1).sub(1).dropna().apply(xl.utils.cell.get_column_letter)

    # Get Column lettters in row 12
    d= {}
    [d.update({x.value:x.column_letter}) for x in r if x.value]

    fromcell = list(map(xl.utils.cell.column_index_from_string,d.values()))

    tocell = [ xl.utils.cell.get_column_letter(fromcell[i+1]-1)
              if  i < len(fromcell)-1 else None
              for i,c in enumerate(fromcell)]

    dd = Series(d,name='from').to_frame()
    # Fill in the 'to' column for the scenario
    dd['to'] = get_to_column(dd['from'])

    # Update scenario dictionary with data range for scenario
    { scenario_dict[k].update({'col_range': [v,tocell[i]] })
     for i,(k,v) in enumerate(d.items()) if k in scenario_dict}

    # Add scenario items and column index range to dictionary
    for scenario_name, scenario_params in scenario_dict.items():
        # Get Item Names
        # List of Column indices [start, end]
        cols = [xl.utils.cell.column_index_from_string(c) for c in scenario_params['col_range']]
        # Use column indices to read row in excel sheet
        item_row = hdr_dict['Item Name']
        iterator = ws.iter_rows(min_col=cols[0], max_col=cols[1],
                                min_row=item_row, max_row=item_row)
        # Extract row
        r = next(iterator)
        # Update scenario dictionary with a list of MIKE items. Screen for "0" items
        scenario_dict[scenario_name].update({'items':[c.value for c in r if c.value.split('|')[0]!='0']})

    #Close Workbook
    wb.close()

    # --------------------------
    # WRITE TO TEMP EXCEL BOOK
    # -------------------------
    # Create Temporary book

    wb = xl.Workbook(write_only=True)
    for scenario_name, scenario_params in scenario_dict.items():
        # Load the results file into Python
        ds = read(scenario_params['Filename'],items = scenario_params['items'])

        # Make data array, transpose
        data = np.transpose(np.asarray(ds.data))

        # Create sheet for this scenario
        ws = wb.create_sheet(title=scenario_name)
        # Row 1: Item names
        ws.append( [None]+[item.name for item in ds.items ] )

        for i in range(len(ds.time)):
            ws.append([ds.time[i]]+data[i,:].tolist())

    # Save once the scenarios are written to sheets
    wb.save(ftemppath)
    wb.close()


if __name__ == "__main__":

    if mode == 'w':
        excel2dfs0()
    elif mode == 'r':
        dfs02excel()
